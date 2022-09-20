# basic

#-*- coding: utf-8 -*-

# need 'pip install'
import json  # pip install json
import pymysql  # pip install pymysql
import requests
from bs4 import BeautifulSoup
from keyring.backends import null
from openpyxl import Workbook
import datetime

currDir = '/Users/KimChangHwi/Desktop/project/telegramBot/InnisfreeChapture/info'


def doCheck():


    urlCheckList = urlCheck()


def urlCheck():

    searchWords = "힙지로"

    textResult = connectCheck(searchWords)

    if textResult == "":
        textResult = "OK"

    return textResult

def connectCheck(searchWords):

    url = 'https://map.naver.com/v5/api/search?caller=pcweb&query={}&type=all&page=1&displayCount=100&isPlaceRecommendationReplace=true&lang=ko'
    url = url.format(searchWords)
    response = requests.get(url, verify=False)
    html = response.text

    write_wb = Workbook()
    write_ws = write_wb.active



    json_object = json.loads(html)
    result = json_object['result']
    place = result['place']

    i =2;
    for list in place['list']:
        write_ws = write_wb.active
        write_ws['A1'] = 'SEARCH_WORDS'
        write_ws['B1'] = 'ID'
        write_ws['C1'] = 'RANK'
        write_ws['D1'] = 'NAME'
        write_ws['E1'] = 'TEL'
        write_ws['F1'] = 'CATEGORY'
        write_ws['G1'] = 'ROADADDRESS'
        write_ws['H1'] = 'CONTEXT'
        write_ws['I1'] = 'THUMURL'
        write_ws['J1'] = 'REVIEWCOUNT'
        write_ws['K1'] = 'HOMEPAGE'
        write_ws['L1'] = 'COLLECT_TIME'
        write_ws.cell(i,1,searchWords);
        j=2;
        id = list['id']
        write_ws.cell(i,j,id); j+=1
        rank = list['rank']
        write_ws.cell(i,j,rank); j+=1
        name = list['name']
        write_ws.cell(i,j,name); j+=1
        tel = list['tel']
        write_ws.cell(i,j,tel); j+=1
        category = list['category']
        category = ', '.join(str(x) for x in category)
        write_ws.cell(i,j,category); j+=1
        roadAddress = list['roadAddress']
        write_ws.cell(i,j,roadAddress); j+=1
        context = list['context']
        context = ', '.join(str(x) for x in context)
        write_ws.cell(i,j,context); j+=1
        thumUrl = list['thumUrl']
        write_ws.cell(i,j,thumUrl); j+=1
        reviewCount = list['reviewCount']
        write_ws.cell(i,j,reviewCount); j+=1
        homePage = list['homePage']
        write_ws.cell(i,j,homePage); j+=1

        now = datetime.datetime.now()
        nowDatetime = now.strftime('%Y-%m-%d %H:%M:%S')
        write_ws.cell(i,j,nowDatetime); j+=1

        i=i+1;

    write_wb.save("/Users/KimChangHwi/Desktop/project/naverplace/src/숫자.xlsx")
    return "textResult"


def makeMessage(urlCheckList):

    message = urlCheckList

    return message

def main():
    # get data from data.json
    with open(currDir + "/" + "data.json", encoding="UTF-8") as json_file:
        json_data = json.load(json_file)

        doCheck()



if __name__ == "__main__":
    main()
